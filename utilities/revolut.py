import openpyxl
import datetime
from utilities import NBP_APIs
from utilities.find_fx_rate import fx
from utilities import user_interactions


def file_ops(file, user_input, rates):
    """
    Calculates tax based on transactions in the provided Excel file.

    Parameters:
    file (str): The path to the Excel file containing the transaction data.
    user_input (str): The year for which the tax calculation is being performed.
    rates (str): Path to Excel file containing exchange rates, from NBP.

    Returns:
    None. Prints the calculated expenses, income, taxable income, and tax to be paid.
    """
    expenses = 0
    income = 0

    revolut_file = openpyxl.load_workbook(file)
    revolut_sheet_origin = revolut_file.active
    sell_sheet = revolut_file.create_sheet(title='sell')
    buy_sheet = revolut_file.create_sheet(title='buy')

    sell_sheet, buy_sheet = append_sheets(user_input, revolut_sheet_origin, sell_sheet, buy_sheet)

    buy_sell = [sell_sheet, buy_sheet]

    for sheet in buy_sell:
        for row in range(1, sheet.max_row + 1):
            date = sheet.cell(row=row, column=1).value.split('T')[0]
            transaction_date = datetime.datetime.strptime(date, '%Y-%m-%d')
            exchange_date = transaction_date - datetime.timedelta(days=1)
            currency = sheet.cell(row=row, column=7).value
            fx_rate = NBP_APIs.api_request(currency, exchange_date)
            # fx_rate = fx(transaction_date, currency, rates, False)
            sheet.cell(row=row, column=9).value = fx_rate
            sheet.cell(row=row, column=10).value = sheet.cell(row=row, column=6).value * fx_rate
            if sheet == sell_sheet:
                income += sheet.cell(row=row, column=10).value
            else:
                expenses += sheet.cell(row=row, column=10).value

    user_interactions.results(expenses, income)
    # revolut_file.save(file)


def append_sheets(user_input, origin_sheet, sell_sheet, buy_sheet):
    """
    Filters and appends relevant transactions to separate sheets based on the transaction type and year.

    Parameters:
    user_input (str): The year for which transactions are being filtered.
    origin_sheet (Worksheet): The original worksheet containing all transaction data.
    sell_sheet (Worksheet): The worksheet where 'SELL - MARKET' transactions will be appended.
    buy_sheet (Worksheet): The worksheet where 'BUY - MARKET' transactions will be appended.

    Returns:
    tuple: A tuple containing the updated sell_sheet and buy_sheet with the relevant transactions appended.
    """

    for row in range(2, origin_sheet.max_row + 1):
        if origin_sheet.cell(row=row, column=1).value[:4].split('T')[0][:4] == user_input and \
                origin_sheet.cell(row=row, column=3).value == 'SELL - MARKET':
            row_to_copy = [cell.value for cell in origin_sheet[row]]
            sell_sheet.append(row_to_copy)
        elif origin_sheet.cell(row=row, column=3).value == 'BUY - MARKET':
            row_to_copy = [cell.value for cell in origin_sheet[row]]
            buy_sheet.append(row_to_copy)

    return sell_sheet, buy_sheet
