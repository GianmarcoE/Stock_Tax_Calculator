import openpyxl
import datetime


def fx(transaction_date, currency, rates, repeat) -> float:
    """
    Retrieves exchange rate from NBP Excel file, in case APIs are down.

    This function retrieves the exchange rate for a given currency on a specified transaction date (-1)
    from an Excel file containing historical exchange rates. If the rate is not found for the given
    date, it recursively searches for the rate on previous dates. This to overcome the issue of local holidays with
    no rates.

    Parameters:
    transaction_date (datetime.date): The date of the transaction for which the exchange rate is needed.
    currency (str): The currency code ('USD' or 'EUR') for which the exchange rate is required.
    rates (str): The file path to the Excel workbook containing the exchange rates.
    repeat (bool): A flag indicating whether to use the rate from the current date (True) or the previous date (False).

    Returns:
    float: The exchange rate for the specified currency on the given transaction date or the closest previous date.
    """

    # Load rates list from NBP
    rates_file = openpyxl.load_workbook(rates)
    rates_sheet = rates_file.active

    exchange_rate = None
    for row in range(3, rates_sheet.max_row + 1):
        ref_date = rates_sheet.cell(row=row, column=1).value
        if ref_date == transaction_date:
            if currency == 'USD' and repeat is False:
                exchange_rate = rates_sheet.cell(row=row-1, column=2).value
            elif currency == 'USD' and repeat is True:
                exchange_rate = rates_sheet.cell(row=row, column=2).value
            elif currency == 'EUR' and repeat is False:
                exchange_rate = rates_sheet.cell(row=row-1, column=3).value
            elif currency == 'EUR' and repeat is True:
                exchange_rate = rates_sheet.cell(row=row, column=3).value
            break

    while exchange_rate is None:
        transaction_date = transaction_date - datetime.timedelta(days=1)
        exchange_rate = fx(transaction_date, currency, rates, True)

    return exchange_rate
